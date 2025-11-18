"""
Ofertas Service for TeLOO V3
Handles individual and bulk offer creation, validation, and state management
"""

import logging
from typing import List, Dict, Optional, Any, Tuple
from decimal import Decimal
from datetime import datetime, timedelta
import json
import pandas as pd
import io

from models.oferta import Oferta, OfertaDetalle
from models.solicitud import Solicitud, RepuestoSolicitado
from models.user import Asesor
from models.enums import EstadoSolicitud, EstadoOferta, OrigenOferta
from services.concurrencia_service import ConcurrenciaService

logger = logging.getLogger(__name__)


class OfertasService:
    """Service for managing ofertas (offers)"""
    
    @staticmethod
    async def create_oferta_individual(
        solicitud_id: str,
        asesor_id: str,
        tiempo_entrega_dias: int,
        observaciones: Optional[str],
        detalles: List[Dict[str, Any]],
        redis_client=None
    ) -> Dict[str, Any]:
        """
        Create individual offer from form data
        
        Args:
            solicitud_id: ID of the solicitud
            asesor_id: ID of the asesor making the offer
            tiempo_entrega_dias: General delivery time for the entire order
            observaciones: Optional observations
            detalles: List of offer details per repuesto
            redis_client: Redis client for publishing events
            
        Returns:
            Dict with created offer data
        """
        from tortoise.transactions import in_transaction
        
        try:
            # Validate solicitud exists and is ABIERTA
            solicitud = await Solicitud.get_or_none(id=solicitud_id)
            if not solicitud:
                raise ValueError(f"Solicitud {solicitud_id} no encontrada")
                
            if solicitud.estado != EstadoSolicitud.ABIERTA:
                raise ValueError(f"Solicitud {solicitud_id} no está en estado ABIERTA")
            
            # Validate concurrency - no evaluation in progress
            await ConcurrenciaService.validar_oferta_concurrencia(solicitud_id, redis_client)
            
            # Validate asesor exists
            asesor = await Asesor.get_or_none(id=asesor_id)
            if not asesor:
                raise ValueError(f"Asesor {asesor_id} no encontrado")
            
            # Validate tiempo_entrega_dias range (0-90)
            if not (0 <= tiempo_entrega_dias <= 90):
                raise ValueError("Tiempo de entrega debe estar entre 0 y 90 días")
            
            # Check if asesor already has an offer for this solicitud
            # Use filter().first() to handle multiple results (cleanup old duplicates)
            oferta_existente = await Oferta.filter(
                solicitud_id=solicitud_id,
                asesor_id=asesor_id
            ).first()
            
            # Calculate expiration date
            from services.configuracion_service import ConfiguracionService
            config = await ConfiguracionService.get_config('parametros_generales')
            timeout_horas = config.get('timeout_ofertas_horas', 20)
            
            # Check if solicitud has custom timeout
            solicitud_timeout = getattr(solicitud, 'timeout_horas', None)
            if solicitud_timeout:
                timeout_horas = solicitud_timeout
            
            from utils.datetime_utils import now_utc, add_hours
            fecha_expiracion = add_hours(now_utc(), timeout_horas)
            
            # Use atomic transaction to ensure data consistency
            async with in_transaction() as conn:
                # If exists, update it instead of creating new
                if oferta_existente:
                    # Update existing offer
                    oferta_existente.tiempo_entrega_dias = tiempo_entrega_dias
                    oferta_existente.observaciones = observaciones
                    oferta_existente.estado = EstadoOferta.ENVIADA
                    oferta_existente.fecha_expiracion = fecha_expiracion
                    oferta_existente.updated_at = now_utc()
                    await oferta_existente.save(using_db=conn)
                    
                    # Delete existing details to replace with new ones
                    await OfertaDetalle.filter(oferta_id=oferta_existente.id).using_db(conn).delete()
                    
                    oferta = oferta_existente
                else:
                    # Create new offer
                    oferta = await Oferta.create(
                        solicitud=solicitud,
                        asesor=asesor,
                        tiempo_entrega_dias=tiempo_entrega_dias,
                        observaciones=observaciones,
                        estado=EstadoOferta.ENVIADA,
                        origen=OrigenOferta.FORM,
                        fecha_expiracion=fecha_expiracion,
                        using_db=conn
                    )
                
                # Process each detail
                monto_total = Decimal('0.00')
                detalles_creados = []
                
                for detalle_data in detalles:
                    # Validate repuesto exists
                    repuesto_id = detalle_data.get('repuesto_solicitado_id')
                    repuesto = await RepuestoSolicitado.get_or_none(
                        id=repuesto_id,
                        solicitud=solicitud
                    )
                    if not repuesto:
                        raise ValueError(f"Repuesto {repuesto_id} no encontrado en la solicitud")
                    
                    # Validate ranges
                    precio_unitario = Decimal(str(detalle_data.get('precio_unitario', 0)))
                    garantia_meses = int(detalle_data.get('garantia_meses', 1))
                    cantidad = int(detalle_data.get('cantidad', 1))
                    
                    # Validate precio range (1000-50000000)
                    if not (1000 <= precio_unitario <= 50000000):
                        raise ValueError(f"Precio unitario debe estar entre 1,000 y 50,000,000 COP")
                    
                    # Validate garantia range (1-60)
                    if not (1 <= garantia_meses <= 60):
                        raise ValueError(f"Garantía debe estar entre 1 y 60 meses")
                    
                    # Create offer detail
                    detalle = await OfertaDetalle.create(
                        oferta=oferta,
                        repuesto_solicitado=repuesto,
                        precio_unitario=precio_unitario,
                        cantidad=cantidad,
                        garantia_meses=garantia_meses,
                        tiempo_entrega_dias=tiempo_entrega_dias,  # Use general delivery time
                        marca_repuesto=detalle_data.get('marca_repuesto'),
                        modelo_repuesto=detalle_data.get('modelo_repuesto'),
                        origen_repuesto=detalle_data.get('origen_repuesto'),
                        observaciones=detalle_data.get('observaciones'),
                        using_db=conn
                    )
                    
                    detalles_creados.append(detalle)
                    monto_total += precio_unitario * cantidad
                
                # Update offer totals
                oferta.monto_total = monto_total
                oferta.cantidad_repuestos = len(detalles_creados)
                oferta.cobertura_porcentaje = await oferta.calcular_cobertura()
                await oferta.save(using_db=conn)
            
            # Transaction committed successfully at this point
            
            # Publish event to Redis
            if redis_client:
                try:
                    evento_data = {
                        'tipo_evento': 'oferta.created',
                        'oferta_id': str(oferta.id),
                        'solicitud_id': str(solicitud.id),
                        'asesor_id': str(asesor.id),
                        'monto_total': float(monto_total),
                        'cantidad_repuestos': len(detalles_creados),
                        'cobertura_porcentaje': float(await oferta.calcular_cobertura()),
                        'origen': 'FORM',
                        'timestamp': datetime.now().isoformat()
                    }
                    
                    await redis_client.publish('oferta.created', json.dumps(evento_data))
                    logger.info(f"Evento oferta.created publicado para oferta {oferta.id}")
                    
                except Exception as e:
                    logger.error(f"Error publicando evento a Redis: {e}")
            
            # Load relationships for response
            await oferta.fetch_related('solicitud', 'asesor__usuario', 'detalles__repuesto_solicitado')
            
            return {
                'success': True,
                'oferta_id': str(oferta.id),
                'codigo_oferta': oferta.codigo_oferta,
                'monto_total': float(oferta.monto_total),
                'cantidad_repuestos': oferta.cantidad_repuestos,
                'cobertura_porcentaje': float(oferta.cobertura_porcentaje),
                'detalles_count': len(detalles_creados),
                'message': f'Oferta creada exitosamente con {len(detalles_creados)} repuestos'
            }
            
        except ValueError as e:
            logger.error(f"Error de validación creando oferta: {e}")
            raise
        except Exception as e:
            logger.error(f"Error inesperado creando oferta: {e}")
            raise ValueError(f"Error interno del servidor: {str(e)}")
    
    @staticmethod
    async def get_oferta_by_id(oferta_id: str) -> Optional[Oferta]:
        """Get offer by ID with all relationships"""
        try:
            oferta = await Oferta.filter(id=oferta_id).prefetch_related(
                'solicitud__repuestos_solicitados',
                'asesor__usuario',
                'detalles__repuesto_solicitado'
            ).first()
            return oferta
        except Exception as e:
            logger.error(f"Error obteniendo oferta {oferta_id}: {e}")
            return None
    
    @staticmethod
    async def get_ofertas_by_solicitud(solicitud_id: str) -> List[Oferta]:
        """Get all offers for a solicitud"""
        try:
            ofertas = await Oferta.filter(solicitud_id=solicitud_id).prefetch_related(
                'asesor__usuario',
                'detalles__repuesto_solicitado'
            ).order_by('-created_at')
            return ofertas
        except Exception as e:
            logger.error(f"Error obteniendo ofertas para solicitud {solicitud_id}: {e}")
            return []
    
    @staticmethod
    async def validate_oferta_data(
        solicitud_id: str,
        asesor_id: str,
        tiempo_entrega_dias: int,
        detalles: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Validate offer data before creation
        
        Returns:
            Dict with validation result and errors if any
        """
        errors = []
        warnings = []
        
        try:
            # Validate solicitud
            solicitud = await Solicitud.get_or_none(id=solicitud_id)
            if not solicitud:
                errors.append(f"Solicitud {solicitud_id} no encontrada")
                return {'valid': False, 'errors': errors, 'warnings': warnings}
            
            if solicitud.estado != EstadoSolicitud.ABIERTA:
                errors.append(f"Solicitud no está en estado ABIERTA (estado actual: {solicitud.estado})")
            
            # Validate asesor
            asesor = await Asesor.get_or_none(id=asesor_id)
            if not asesor:
                errors.append(f"Asesor {asesor_id} no encontrado")
            
            # Validate tiempo_entrega_dias
            if not (0 <= tiempo_entrega_dias <= 90):
                errors.append("Tiempo de entrega debe estar entre 0 y 90 días")
            
            # Validate detalles
            if not detalles:
                errors.append("Debe incluir al menos un repuesto en la oferta")
            
            repuestos_solicitud = await RepuestoSolicitado.filter(solicitud=solicitud)
            repuestos_ids = {str(r.id) for r in repuestos_solicitud}
            
            for i, detalle in enumerate(detalles):
                detalle_errors = []
                
                # Validate repuesto exists
                repuesto_id = detalle.get('repuesto_solicitado_id')
                if not repuesto_id:
                    detalle_errors.append("repuesto_solicitado_id es requerido")
                elif str(repuesto_id) not in repuestos_ids:
                    detalle_errors.append(f"Repuesto {repuesto_id} no pertenece a la solicitud")
                
                # Validate precio
                try:
                    precio = Decimal(str(detalle.get('precio_unitario', 0)))
                    if not (1000 <= precio <= 50000000):
                        detalle_errors.append("Precio debe estar entre 1,000 y 50,000,000 COP")
                except (ValueError, TypeError):
                    detalle_errors.append("Precio unitario debe ser un número válido")
                
                # Validate garantia
                try:
                    garantia = int(detalle.get('garantia_meses', 1))
                    if not (1 <= garantia <= 60):
                        detalle_errors.append("Garantía debe estar entre 1 y 60 meses")
                except (ValueError, TypeError):
                    detalle_errors.append("Garantía debe ser un número entero válido")
                
                # Validate cantidad
                try:
                    cantidad = int(detalle.get('cantidad', 1))
                    if cantidad < 1:
                        detalle_errors.append("Cantidad debe ser mayor a 0")
                except (ValueError, TypeError):
                    detalle_errors.append("Cantidad debe ser un número entero válido")
                
                if detalle_errors:
                    errors.extend([f"Detalle {i+1}: {error}" for error in detalle_errors])
            
            # Check for duplicate repuestos
            repuestos_en_oferta = [d.get('repuesto_solicitado_id') for d in detalles]
            duplicados = set([r for r in repuestos_en_oferta if repuestos_en_oferta.count(r) > 1])
            if duplicados:
                errors.append(f"Repuestos duplicados en la oferta: {duplicados}")
            
            # Calculate coverage
            if repuestos_solicitud:
                cobertura = (len(detalles) / len(repuestos_solicitud)) * 100
                if cobertura < 50:
                    warnings.append(f"Cobertura baja: {cobertura:.1f}% (recomendado ≥50%)")
            
            return {
                'valid': len(errors) == 0,
                'errors': errors,
                'warnings': warnings,
                'cobertura_estimada': cobertura if repuestos_solicitud else 0
            }
            
        except Exception as e:
            logger.error(f"Error validando datos de oferta: {e}")
            return {
                'valid': False,
                'errors': [f"Error interno de validación: {str(e)}"],
                'warnings': []
            }
    
    @staticmethod
    async def create_oferta_bulk_excel(
        solicitud_id: str,
        asesor_id: str,
        excel_file_content: bytes,
        filename: str,
        redis_client=None
    ) -> Dict[str, Any]:
        """
        Create offer from Excel file upload
        
        Args:
            solicitud_id: ID of the solicitud
            asesor_id: ID of the asesor making the offer
            excel_file_content: Excel file content as bytes
            filename: Original filename for logging
            redis_client: Redis client for publishing events
            
        Returns:
            Dict with creation result and validation errors
        """
        try:
            # Validate concurrency first - no evaluation in progress
            await ConcurrenciaService.validar_oferta_concurrencia(solicitud_id, redis_client)
            
            # Validate file format and size
            if not filename.lower().endswith('.xlsx'):
                raise ValueError("Archivo debe ser formato .xlsx")
            
            # Validate file size (5MB max)
            max_size = 5 * 1024 * 1024  # 5MB
            if len(excel_file_content) > max_size:
                raise ValueError(f"Archivo excede el tamaño máximo de 5MB (actual: {len(excel_file_content)/1024/1024:.1f}MB)")
            
            # Parse Excel file
            validation_result = await OfertasService.parse_and_validate_excel(
                excel_file_content, solicitud_id, asesor_id
            )
            
            if not validation_result['valid']:
                return {
                    'success': False,
                    'errors': validation_result['errors'],
                    'warnings': validation_result.get('warnings', []),
                    'rows_processed': validation_result.get('rows_processed', 0),
                    'message': 'Validación de Excel falló'
                }
            
            # Extract validated data
            tiempo_entrega_dias = validation_result['tiempo_entrega_dias']
            observaciones = validation_result['observaciones']
            detalles = validation_result['detalles']
            
            # Create offer using individual method
            result = await OfertasService.create_oferta_individual(
                solicitud_id=solicitud_id,
                asesor_id=asesor_id,
                tiempo_entrega_dias=tiempo_entrega_dias,
                observaciones=observaciones,
                detalles=detalles,
                redis_client=None  # Will publish bulk event instead
            )
            
            # Publish bulk upload event to Redis
            if redis_client:
                try:
                    evento_data = {
                        'tipo_evento': 'oferta.bulk_uploaded',
                        'oferta_id': result['oferta_id'],
                        'solicitud_id': solicitud_id,
                        'asesor_id': asesor_id,
                        'filename': filename,
                        'rows_processed': validation_result['rows_processed'],
                        'monto_total': result['monto_total'],
                        'cantidad_repuestos': result['cantidad_repuestos'],
                        'cobertura_porcentaje': result['cobertura_porcentaje'],
                        'origen': 'EXCEL',
                        'timestamp': datetime.now().isoformat()
                    }
                    
                    await redis_client.publish('oferta.bulk_uploaded', json.dumps(evento_data))
                    logger.info(f"Evento oferta.bulk_uploaded publicado para oferta {result['oferta_id']}")
                    
                except Exception as e:
                    logger.error(f"Error publicando evento bulk a Redis: {e}")
            
            return {
                'success': True,
                'oferta_id': result['oferta_id'],
                'codigo_oferta': result['codigo_oferta'],
                'monto_total': result['monto_total'],
                'cantidad_repuestos': result['cantidad_repuestos'],
                'cobertura_porcentaje': result['cobertura_porcentaje'],
                'rows_processed': validation_result['rows_processed'],
                'warnings': validation_result.get('warnings', []),
                'message': f'Oferta creada exitosamente desde Excel con {result["cantidad_repuestos"]} repuestos'
            }
            
        except ValueError as e:
            logger.error(f"Error de validación en bulk upload: {e}")
            raise
        except Exception as e:
            logger.error(f"Error inesperado en bulk upload: {e}")
            raise ValueError(f"Error procesando archivo Excel: {str(e)}")
    
    @staticmethod
    async def parse_and_validate_excel(
        excel_content: bytes,
        solicitud_id: str,
        asesor_id: str
    ) -> Dict[str, Any]:
        """
        Parse and validate Excel file content
        
        Expected columns:
        - repuesto_nombre: Name of the repuesto (for matching)
        - precio_unitario: Unit price (1000-50000000)
        - cantidad: Quantity (default 1)
        - garantia_meses: Warranty in months (1-60)
        - marca_repuesto: Optional brand
        - modelo_repuesto: Optional model
        - origen_repuesto: Optional origin
        - observaciones: Optional observations
        
        First row should contain general info:
        - tiempo_entrega_dias: General delivery time
        - observaciones_generales: General observations
        
        Returns:
            Dict with validation result and parsed data
        """
        errors = []
        warnings = []
        detalles = []
        rows_processed = 0
        
        try:
            # Read Excel file
            excel_file = io.BytesIO(excel_content)
            df = pd.read_excel(excel_file, engine='openpyxl')
            
            if df.empty:
                return {
                    'valid': False,
                    'errors': ['Archivo Excel está vacío'],
                    'warnings': warnings,
                    'rows_processed': 0
                }
            
            # Validate required columns
            required_columns = ['repuesto_nombre', 'precio_unitario', 'garantia_meses']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return {
                    'valid': False,
                    'errors': [f'Columnas requeridas faltantes: {missing_columns}'],
                    'warnings': warnings,
                    'rows_processed': 0
                }
            
            # Get solicitud and repuestos
            solicitud = await Solicitud.get_or_none(id=solicitud_id)
            if not solicitud:
                return {
                    'valid': False,
                    'errors': [f'Solicitud {solicitud_id} no encontrada'],
                    'warnings': warnings,
                    'rows_processed': 0
                }
            
            repuestos_solicitud = await RepuestoSolicitado.filter(solicitud=solicitud)
            repuestos_map = {r.nombre.lower().strip(): r for r in repuestos_solicitud}
            
            # Extract general info from first row or use defaults
            tiempo_entrega_dias = 1  # Default
            observaciones_generales = None
            
            if 'tiempo_entrega_dias' in df.columns and not pd.isna(df.iloc[0]['tiempo_entrega_dias']):
                try:
                    tiempo_entrega_dias = int(df.iloc[0]['tiempo_entrega_dias'])
                    if not (0 <= tiempo_entrega_dias <= 90):
                        errors.append("Tiempo de entrega debe estar entre 0 y 90 días")
                except (ValueError, TypeError):
                    errors.append("Tiempo de entrega debe ser un número entero")
            
            if 'observaciones_generales' in df.columns and not pd.isna(df.iloc[0]['observaciones_generales']):
                observaciones_generales = str(df.iloc[0]['observaciones_generales'])
            
            # Process each row
            for index, row in df.iterrows():
                rows_processed += 1
                row_errors = []
                
                # Skip empty rows
                if pd.isna(row['repuesto_nombre']) or str(row['repuesto_nombre']).strip() == '':
                    continue
                
                # Match repuesto by name
                repuesto_nombre = str(row['repuesto_nombre']).lower().strip()
                repuesto = repuestos_map.get(repuesto_nombre)
                
                if not repuesto:
                    # Try partial matching
                    matches = [r for name, r in repuestos_map.items() if repuesto_nombre in name or name in repuesto_nombre]
                    if len(matches) == 1:
                        repuesto = matches[0]
                        warnings.append(f"Fila {index+1}: Coincidencia parcial para '{row['repuesto_nombre']}' → '{repuesto.nombre}'")
                    else:
                        row_errors.append(f"Repuesto '{row['repuesto_nombre']}' no encontrado en la solicitud")
                
                # Validate precio_unitario
                try:
                    precio_unitario = Decimal(str(row['precio_unitario']))
                    if not (1000 <= precio_unitario <= 50000000):
                        row_errors.append("Precio debe estar entre 1,000 y 50,000,000 COP")
                except (ValueError, TypeError):
                    row_errors.append("Precio unitario debe ser un número válido")
                    precio_unitario = Decimal('0')
                
                # Validate garantia_meses
                try:
                    garantia_meses = int(row['garantia_meses'])
                    if not (1 <= garantia_meses <= 60):
                        row_errors.append("Garantía debe estar entre 1 y 60 meses")
                except (ValueError, TypeError):
                    row_errors.append("Garantía debe ser un número entero válido")
                    garantia_meses = 1
                
                # Validate cantidad (optional, default 1)
                cantidad = 1
                if 'cantidad' in df.columns and not pd.isna(row['cantidad']):
                    try:
                        cantidad = int(row['cantidad'])
                        if cantidad < 1:
                            row_errors.append("Cantidad debe ser mayor a 0")
                    except (ValueError, TypeError):
                        row_errors.append("Cantidad debe ser un número entero válido")
                
                # Add row errors to main errors list
                if row_errors:
                    errors.extend([f"Fila {index+1}: {error}" for error in row_errors])
                    continue
                
                # Create detail dict if no errors
                if repuesto:
                    detalle = {
                        'repuesto_solicitado_id': str(repuesto.id),
                        'precio_unitario': precio_unitario,
                        'cantidad': cantidad,
                        'garantia_meses': garantia_meses,
                        'marca_repuesto': str(row.get('marca_repuesto', '')).strip() or None,
                        'modelo_repuesto': str(row.get('modelo_repuesto', '')).strip() or None,
                        'origen_repuesto': str(row.get('origen_repuesto', '')).strip() or None,
                        'observaciones': str(row.get('observaciones', '')).strip() or None
                    }
                    detalles.append(detalle)
            
            # Check if any valid details were found
            if not detalles and not errors:
                errors.append("No se encontraron repuestos válidos en el archivo")
            
            # Check for duplicates
            repuestos_ids = [d['repuesto_solicitado_id'] for d in detalles]
            duplicados = set([r for r in repuestos_ids if repuestos_ids.count(r) > 1])
            if duplicados:
                errors.append(f"Repuestos duplicados en el archivo: {len(duplicados)} repuestos")
            
            # Calculate coverage
            cobertura = 0
            if repuestos_solicitud:
                cobertura = (len(detalles) / len(repuestos_solicitud)) * 100
                if cobertura < 50:
                    warnings.append(f"Cobertura baja: {cobertura:.1f}% (recomendado ≥50%)")
            
            return {
                'valid': len(errors) == 0,
                'errors': errors,
                'warnings': warnings,
                'rows_processed': rows_processed,
                'tiempo_entrega_dias': tiempo_entrega_dias,
                'observaciones': observaciones_generales,
                'detalles': detalles,
                'cobertura_estimada': cobertura
            }
            
        except Exception as e:
            logger.error(f"Error parseando Excel: {e}")
            return {
                'valid': False,
                'errors': [f"Error procesando archivo Excel: {str(e)}"],
                'warnings': warnings,
                'rows_processed': rows_processed
            }
    
    @staticmethod
    async def actualizar_estado_oferta(
        oferta_id: str,
        nuevo_estado: EstadoOferta,
        motivo: Optional[str] = None,
        redis_client=None
    ) -> Dict[str, Any]:
        """
        Update offer state with validation and audit logging
        
        Args:
            oferta_id: ID of the offer to update
            nuevo_estado: New state to set
            motivo: Optional reason for state change
            redis_client: Redis client for publishing events
            
        Returns:
            Dict with update result
        """
        try:
            # Get offer with relationships
            oferta = await Oferta.filter(id=oferta_id).prefetch_related(
                'solicitud', 'asesor__usuario'
            ).first()
            if not oferta:
                raise ValueError(f"Oferta {oferta_id} no encontrada")
            
            estado_anterior = oferta.estado
            
            # Validate state transition
            transiciones_permitidas = OfertasService._get_transiciones_permitidas(estado_anterior)
            if nuevo_estado not in transiciones_permitidas:
                raise ValueError(
                    f"Transición no permitida: {estado_anterior} → {nuevo_estado}. "
                    f"Transiciones válidas: {transiciones_permitidas}"
                )
            
            # Update state
            await oferta.update_from_dict({
                'estado': nuevo_estado,
                'updated_at': datetime.now()
            })
            
            # Log audit entry
            await OfertasService._log_cambio_estado(
                oferta=oferta,
                estado_anterior=estado_anterior,
                estado_nuevo=nuevo_estado,
                motivo=motivo
            )
            
            # Publish event to Redis
            if redis_client:
                try:
                    evento_data = {
                        'tipo_evento': 'oferta.estado_changed',
                        'oferta_id': str(oferta.id),
                        'solicitud_id': str(oferta.solicitud.id),
                        'asesor_id': str(oferta.asesor.id),
                        'estado_anterior': estado_anterior,
                        'estado_nuevo': nuevo_estado,
                        'motivo': motivo,
                        'timestamp': datetime.now().isoformat()
                    }
                    
                    await redis_client.publish('oferta.estado_changed', json.dumps(evento_data))
                    logger.info(f"Evento oferta.estado_changed publicado para oferta {oferta.id}")
                    
                except Exception as e:
                    logger.error(f"Error publicando evento de cambio de estado a Redis: {e}")
            
            return {
                'success': True,
                'oferta_id': str(oferta.id),
                'estado_anterior': estado_anterior,
                'estado_nuevo': nuevo_estado,
                'motivo': motivo,
                'message': f'Estado actualizado de {estado_anterior} a {nuevo_estado}'
            }
            
        except ValueError as e:
            logger.error(f"Error de validación actualizando estado: {e}")
            raise
        except Exception as e:
            logger.error(f"Error inesperado actualizando estado: {e}")
            raise ValueError(f"Error interno del servidor: {str(e)}")
    
    @staticmethod
    def _get_transiciones_permitidas(estado_actual: EstadoOferta) -> List[EstadoOferta]:
        """
        Get allowed state transitions based on current state
        
        State transition rules (simplified):
        - ENVIADA → GANADORA, NO_SELECCIONADA, EXPIRADA, RECHAZADA
        - GANADORA → ACEPTADA, RECHAZADA, EXPIRADA
        - NO_SELECCIONADA → (final state)
        - ACEPTADA → (final state)
        - RECHAZADA → (final state)
        - EXPIRADA → (final state)
        """
        transiciones = {
            EstadoOferta.ENVIADA: [
                EstadoOferta.GANADORA,
                EstadoOferta.NO_SELECCIONADA,
                EstadoOferta.EXPIRADA,
                EstadoOferta.RECHAZADA
            ],
            EstadoOferta.GANADORA: [
                EstadoOferta.ACEPTADA,
                EstadoOferta.RECHAZADA,
                EstadoOferta.EXPIRADA
            ],
            EstadoOferta.NO_SELECCIONADA: [],  # Final state
            EstadoOferta.ACEPTADA: [],  # Final state
            EstadoOferta.RECHAZADA: [],  # Final state
            EstadoOferta.EXPIRADA: []  # Final state
        }
        
        return transiciones.get(estado_actual, [])
    
    @staticmethod
    async def _log_cambio_estado(
        oferta: Oferta,
        estado_anterior: EstadoOferta,
        estado_nuevo: EstadoOferta,
        motivo: Optional[str]
    ):
        """
        Log state change for audit purposes
        
        This would typically write to an audit log table or external logging system
        For now, we'll use structured logging
        """
        try:
            # Import here to avoid circular imports
            from models.analytics import LogAuditoria
            
            # Create audit log entry
            await LogAuditoria.create(
                actor_id=None,  # System action, could be user_id if available
                accion='cambio_estado_oferta',
                entidad='Oferta',
                entidad_id=oferta.id,
                diff_json={
                    'estado_anterior': estado_anterior,
                    'estado_nuevo': estado_nuevo,
                    'motivo': motivo,
                    'solicitud_id': str(oferta.solicitud.id),
                    'asesor_id': str(oferta.asesor.id),
                    'codigo_oferta': oferta.codigo_oferta
                }
            )
            
            logger.info(
                f"Estado de oferta actualizado",
                extra={
                    'oferta_id': str(oferta.id),
                    'codigo_oferta': oferta.codigo_oferta,
                    'solicitud_id': str(oferta.solicitud.id),
                    'asesor_id': str(oferta.asesor.id),
                    'estado_anterior': estado_anterior,
                    'estado_nuevo': estado_nuevo,
                    'motivo': motivo
                }
            )
            
        except Exception as e:
            logger.error(f"Error registrando cambio de estado en auditoría: {e}")
            # Don't raise - audit logging failure shouldn't break the main operation
    
    @staticmethod
    async def get_ofertas_by_estado(estado: EstadoOferta) -> List[Oferta]:
        """Get all offers by state"""
        try:
            ofertas = await Oferta.filter(estado=estado).prefetch_related(
                'solicitud',
                'asesor__usuario',
                'detalles__repuesto_solicitado'
            ).order_by('-created_at')
            return ofertas
        except Exception as e:
            logger.error(f"Error obteniendo ofertas por estado {estado}: {e}")
            return []
    
    @staticmethod
    async def marcar_ofertas_expiradas(
        horas_expiracion: int = 20,
        redis_client=None
    ) -> Dict[str, Any]:
        """
        Mark offers as expired after specified hours
        
        Args:
            horas_expiracion: Hours after which offers expire (default 20)
            redis_client: Redis client for publishing events
            
        Returns:
            Dict with expiration results
        """
        try:
            from datetime import timedelta
            
            # Calculate expiration cutoff
            from utils.datetime_utils import now_utc, add_hours
            cutoff_time = add_hours(now_utc(), -horas_expiracion)
            
            # Find offers to expire
            ofertas_a_expirar = await Oferta.filter(
                estado=EstadoOferta.ENVIADA,
                created_at__lt=cutoff_time
            ).prefetch_related('solicitud', 'asesor__usuario')
            
            ofertas_expiradas = []
            
            for oferta in ofertas_a_expirar:
                try:
                    # Update state
                    await OfertasService.actualizar_estado_oferta(
                        oferta_id=str(oferta.id),
                        nuevo_estado=EstadoOferta.EXPIRADA,
                        motivo=f"Expiración automática después de {horas_expiracion} horas",
                        redis_client=redis_client
                    )
                    ofertas_expiradas.append(str(oferta.id))
                    
                except Exception as e:
                    logger.error(f"Error expirando oferta {oferta.id}: {e}")
            
            logger.info(f"Proceso de expiración completado: {len(ofertas_expiradas)} ofertas expiradas")
            
            return {
                'success': True,
                'ofertas_expiradas': len(ofertas_expiradas),
                'ofertas_ids': ofertas_expiradas,
                'horas_expiracion': horas_expiracion,
                'message': f'{len(ofertas_expiradas)} ofertas marcadas como expiradas'
            }
            
        except Exception as e:
            logger.error(f"Error en proceso de expiración de ofertas: {e}")
            raise ValueError(f"Error procesando expiración: {str(e)}")
    
    @staticmethod
    def generate_excel_template(repuestos_solicitud: List[RepuestoSolicitado]) -> bytes:
        """
        Generate Excel template for bulk offer upload
        
        Args:
            repuestos_solicitud: List of repuestos from the solicitud
            
        Returns:
            Excel file content as bytes
        """
        try:
            # Create template data
            template_data = []
            
            # Add header row with general info
            template_data.append({
                'repuesto_nombre': 'CONFIGURACIÓN GENERAL',
                'precio_unitario': '',
                'cantidad': '',
                'garantia_meses': '',
                'marca_repuesto': '',
                'modelo_repuesto': '',
                'origen_repuesto': '',
                'observaciones': '',
                'tiempo_entrega_dias': 1,
                'observaciones_generales': 'Observaciones generales de la oferta'
            })
            
            # Add empty row
            template_data.append({col: '' for col in template_data[0].keys()})
            
            # Add repuestos from solicitud
            for repuesto in repuestos_solicitud:
                template_data.append({
                    'repuesto_nombre': repuesto.nombre,
                    'precio_unitario': '',  # To be filled by asesor
                    'cantidad': repuesto.cantidad,
                    'garantia_meses': '',  # To be filled by asesor
                    'marca_repuesto': '',
                    'modelo_repuesto': '',
                    'origen_repuesto': '',
                    'observaciones': f'Para {repuesto.vehiculo_completo}',
                    'tiempo_entrega_dias': '',
                    'observaciones_generales': ''
                })
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Create Excel file in memory
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Oferta', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Oferta']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Format config row
                config_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=2, column=col_num)
                    cell.fill = config_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generando template Excel: {e}")
            raise ValueError(f"Error generando template: {str(e)}")